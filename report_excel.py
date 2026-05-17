"""
Excel report generation - exports a completed analysis to an .xlsx workbook
with Summary, Bill of Quantities, and S-Curve sheets (including a line chart).
"""

import io

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)
BOLD = Font(bold=True)


def _style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def build_excel_report(report, output=None):
    """Build the Excel workbook from an analysis report dict.

    Returns the workbook as bytes when output is None, otherwise writes to the
    given path / file-like object and returns it.
    """
    wb = Workbook()
    boq = report.get("boq") or {}
    currency = boq.get("currency", "PKR")
    analysis = report.get("analysis", {})

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Construction Volume Analysis Report"
    ws["A1"].font = TITLE_FONT
    summary_rows = [
        ("Project", report.get("project_name", "")),
        ("Task ID", report.get("task_id", "")),
        ("Timestamp", report.get("timestamp", "")),
        ("Net volume change (m3)", report.get("volume_change_m3")),
        ("Cut volume (m3)", analysis.get("cut_volume")),
        ("Fill volume (m3)", analysis.get("fill_volume")),
        ("Interpretation", analysis.get("interpretation", "")),
        (f"Total cost ({currency})", boq.get("total_cost")),
    ]
    row = 3
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=2, value=value)
        row += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45

    # --- Bill of Quantities sheet ---
    ws = wb.create_sheet("Bill of Quantities")
    headers = ["Road Layer", "Quantity", "Unit",
               f"Rate ({currency})", f"Amount ({currency})"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for item in boq.get("items", []):
        ws.append([item["layer"], item["quantity"], item["unit"],
                   item["rate"], item["amount"]])
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=4, value="TOTAL").font = BOLD
    ws.cell(row=total_row, column=5, value=boq.get("total_cost", 0)).font = BOLD
    for col, width in zip("ABCDE", (28, 14, 10, 16, 18)):
        ws.column_dimensions[col].width = width

    # --- S-Curve sheet ---
    rows = (report.get("scurve") or {}).get("rows", [])
    if rows:
        ws = wb.create_sheet("S-Curve")
        ws.append(["Period", "Planned", "Actual", "Variance"])
        _style_header(ws, 1, 4)
        for r in rows:
            ws.append([r["period"], r["planned"], r["actual"], r["variance"]])
        for col in "ABCD":
            ws.column_dimensions[col].width = 16

        chart = LineChart()
        chart.title = "Project S-Curve - Planned vs Actual"
        chart.y_axis.title = f"Cumulative Cost ({currency})"
        chart.x_axis.title = "Period"
        data = Reference(ws, min_col=2, max_col=3, min_row=1,
                         max_row=len(rows) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height, chart.width = 10, 18
        ws.add_chart(chart, "F2")

    # --- output ---
    if output is None:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    wb.save(output)
    return output
